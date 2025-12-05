import openpyxl
from collections import defaultdict
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox


def add_indexes_to_duplicates(file_path):
    # Открываем workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Работаем с активным листом

    # Получаем максимальное количество строк и столбцов
    max_row = ws.max_row
    max_col = ws.max_column

    # Проходим по каждому столбцу
    for col in range(1, max_col + 1):
        # Словарь для подсчета вхождений значений в столбце (начиная со второй строки)
        value_count = defaultdict(int)
        values = []

        # Собираем значения из столбца (со второй строки)
        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=col)
            values.append((row, cell.value))  # Запоминаем строку и значение

        # Подсчитываем вхождения
        for _, val in values:
            if val is not None:  # Игнорируем пустые ячейки
                value_count[val] += 1

        # Теперь проходим по значениям и модифицируем дубликаты
        seen = defaultdict(int)  # Сколько раз мы уже видели это значение
        for row, val in values:
            if val is not None and value_count[val] > 1:
                seen[val] += 1
                if seen[val] > 1:  # Для первого вхождения оставляем как есть, для остальных добавляем индекс
                    new_val = f"{val}_{seen[val] - 1}"
                    ws.cell(row=row, column=col).value = new_val

    # Сохраняем изменения
    wb.save(file_path)
    messagebox.showinfo("Успех", f"Файл '{file_path}' обработан успешно.")


def select_file():
    root = tk.Tk()
    root.withdraw()  # Скрываем основное окно Tkinter
    file_path = filedialog.askopenfilename(
        title="Выберите Excel-файл",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
    )
    if file_path:
        add_indexes_to_duplicates(file_path)
    else:
        messagebox.showwarning("Предупреждение", "Файл не выбран.")


if __name__ == "__main__":
    select_file()
